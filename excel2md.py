#!/usr/bin/env python3
"""将目录下所有 Excel 的每个 sheet 转为独立 markdown 文件，并生成索引。

用法:
    python excel2md.py <目录路径> [--output <输出目录>]

输出:
    <output>/
        <excel_name>__<sheet_name>.md   # 每个 sheet 一个文件
        INDEX.md                        # 索引文件
"""

import argparse
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("错误: 需要 openpyxl，请运行: pip install openpyxl")


EXCEL_EXTS = {".xlsx", ".xls"}
ESCAPE_MAP = str.maketrans({"/": "／", "\\": "＼", ":": "：", "*": "＊",
                            "?": "？", '"': "＂", "<": "＜", ">": "＞",
                            "|": "｜", "\n": " ", "\r": " "})

def safe_path(name: str) -> str:
    """将路径/文件名中的空格替换为下划线，并转义非法字符。"""
    return str(name).replace(" ", "_").translate(ESCAPE_MAP).strip()


def safe_name(name: str) -> str:
    """将文件名中的非法/危险字符转义为全角字符。"""
    return str(name).translate(ESCAPE_MAP).strip()


def extract_version(path: str) -> str:
    """从路径中提取版本号，例如 feat-1.4.5-main -> 1.4.5。"""
    m = re.search(r"(\d+\.\d+\.\d+)", path)
    return m.group(1) if m else "unknown"


# 排除的"无色"值（openpyxl 中表示无填充）
_NO_FILL = {None, "", "00000000", "00FFFFFF", "FFFFFFFF", "FFFFFF", "000000"}


def _is_highlighted(cell) -> bool:
    """判断单元格是否有非默认的背景色填充。"""
    fill = cell.fill
    if fill.patternType != "solid":
        return False
    raw = fill.start_color.rgb
    if raw is None:
        return False
    raw = str(raw).upper().strip()
    if raw in _NO_FILL:
        return False
    return True


def _cell_text(cell) -> str:
    """获取单元格文本：高亮→反引号，加粗→**加粗**。"""
    val = cell.value
    if val is None:
        return ""
    text = str(val).strip()
    if not text:
        return ""
    text = _clean_format(text)
    if _is_highlighted(cell):
        return f"`{text}`"
    if cell.font and cell.font.bold:
        return f"**{text}**"
    return text


# 页头/封面 sheet 名称匹配模式
_HEADER_SHEET_PATTERNS = [
    r"^头页",
    r"^页头",
    r"^封面",
]


def _is_header_sheet(sheet_name: str) -> bool:
    """判断是否为页头/封面信息 sheet（纯封面元数据，无实际业务内容）。"""
    return any(re.search(p, sheet_name) for p in _HEADER_SHEET_PATTERNS)


def _is_lang_sheet(sheet_name: str) -> bool:
    """判断是否为多语言 sheet，需要保留表格格式输出。"""
    return sheet_name == "多语言"


def _render_table(ws, all_rows) -> str:
    """将 sheet 渲染为 markdown 表格，移除全空行和全空列。"""
    # 获取表头
    header_row = all_rows[0]
    header = [_cell_text(c) for c in header_row]

    # 找出有数据的列（表头非空或下方行有数据）
    data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    valid_cols = []
    for col_idx in range(len(header)):
        # 表头非空，或者该列有任意数据
        has_data = header[col_idx].strip() != ""
        if not has_data:
            for dr in data_rows:
                if col_idx < len(dr) and dr[col_idx] is not None and str(dr[col_idx]).strip():
                    has_data = True
                    break
        if has_data:
            valid_cols.append(col_idx)

    # 渲染表头
    lines = [f"## {ws.title}\n"]
    filtered_header = [header[i] for i in valid_cols]
    lines.append("| " + " | ".join(filtered_header) + " |")
    lines.append("| " + " | ".join("---" for _ in filtered_header) + " |")

    # 渲染数据行，跳过全空行
    data_rows_cells = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False))
    for row in data_rows_cells:
        cells = [_cell_text(c) for c in row]
        while len(cells) < len(header):
            cells.append("")
        cells = cells[:len(header)]

        filtered_cells = [cells[i].strip() for i in valid_cols]
        # 跳过全空行
        if all(c == "" for c in filtered_cells):
            continue

        lines.append("| " + " | ".join(filtered_cells) + " |")

    lines.append("")
    return "\n".join(lines)


def _clean_format(s: str) -> str:
    """去除用户做的格式注释。"""
    return re.sub(r"[（(]加粗样式[)）]$", "", s).strip()


def _heading_level(font_size) -> int:
    """根据加粗字号返回对应的 Markdown header 级别。
    返回: 1→#, 2→##, 3→###, 4→####；0 表示非标题
    """
    if font_size is None:
        return 0
    if font_size >= 16:
        return 1
    if font_size >= 14:
        return 2
    if font_size >= 12:
        return 3
    if font_size >= 11:
        return 4
    return 0


def _outline_level(text: str) -> tuple[int, str]:
    """根据 · 前缀和前导空格判断大纲层级。
    返回 (缩进层级, 清理后的文本)
    - 无 · 前缀 → -1 （作为标题，级别由 _heading_level 决定）
    - · 前无缩进 → 0  （一级列表 `- `）
    - · 前有 2 空格 → 1（二级列表 `  - `）
    - · 前有 4 空格 → 2（三级列表 `    - `）
    """
    stripped = text.lstrip()

    if not stripped.startswith("·"):
        return -1, _clean_format(text)

    lead_spaces = len(text) - len(stripped)
    indent_level = lead_spaces // 2
    cleaned = stripped[1:].strip()  # 去掉 · 和前后空白
    return indent_level, _clean_format(cleaned)


def table_to_md(ws) -> str:
    """将 worksheet 渲染为标题+缩进结构，高亮单元格用反引号标记。

    三种模式：
    1. 多语言 sheet → 保留 markdown 表格格式
    2. 单列表格（仅 1 列有表头）→ 大纲文档模式
    3. 多列表格（≥2 列有表头）→ 表格转条目模式
    """
    all_rows = list(ws.iter_rows(values_only=False))
    if not all_rows:
        return f"## {ws.title}\n\n*空 sheet*\n"

    # === 多语言 sheet：保留表格格式 ===
    if _is_lang_sheet(ws.title):
        return _render_table(ws, all_rows)

    lines = [f"## {ws.title}\n"]

    # 表头列名
    header = [_cell_text(c) for c in all_rows[0]]
    valid_cols = [i for i, h in enumerate(header) if h.strip()]

    if len(valid_cols) <= 1:
        # === 大纲文档模式 ===
        # 先用原始文本判断缩进层级，再应用样式
        items = []  # (row_idx, level, text, hdr_level)
        sheet_title_lower = ws.title.lower()
        for row_idx, row in enumerate(all_rows, start=1):
            for cell in row:
                if cell.value is None:
                    continue
                raw = str(cell.value)
                # 保留原始前导空格用于缩进检测
                level, cleaned = _outline_level(raw)
                if not cleaned:
                    continue
                # 跳过与 sheet 名称相同的首行标题（已在 ## {ws.title} 中输出）
                cleaned_check = _clean_format(raw).strip()
                if row_idx == 1 and cleaned_check.lower() == sheet_title_lower:
                    continue
                # 判断是否为标题（加粗+有字号）
                hdr_level = 0
                if cell.font and cell.font.bold:
                    hdr_level = _heading_level(cell.font.size)
                    if hdr_level > 0:
                        cleaned = _clean_format(cleaned)
                        # 高亮优先级高于加粗
                        if _is_highlighted(cell):
                            cleaned = f"`{cleaned}`"
                        items.append((row_idx, level, cleaned, hdr_level))
                        continue
                # 非标题：应用高亮
                if _is_highlighted(cell):
                    cleaned = f"`{cleaned}`"
                items.append((row_idx, level, cleaned, 0))

        if not items:
            lines.append("*空数据*\n")
        else:
            for _, level, text, hdr_level in items:
                if hdr_level > 0:
                    # 加粗 → 根据字号映射 header 级别
                    lines.append(f"{'#' * hdr_level} {text}")
                elif level == -1:
                    # 无 · 前缀且不加粗 → 默认三级标题
                    lines.append(f"### {text}")
                else:
                    # 有 · 前缀 → 列表项，缩进由 level 决定
                    indent = "  " * level
                    lines.append(f"{indent}- {text}")
            lines.append("")
    else:
        # === 表格转条目模式 ===
        def _row_has_data(cells: list[str]) -> bool:
            for i in valid_cols:
                if i < len(cells) and cells[i].strip():
                    return True
            return False

        for row_idx, row in enumerate(all_rows[1:], start=1):
            # 用 _cell_text 处理高亮/加粗/注释
            cells = [_cell_text(c) for c in row]
            while len(cells) < len(header):
                cells.append("")
            cells = cells[:len(header)]

            if not _row_has_data(cells):
                continue

            title_val = cells[0].strip() if cells[0].strip() else None
            title = f"### {title_val}" if title_val else f"### 行 {row_idx}"
            lines.append(title)

            for i in valid_cols:
                if i == 0:
                    continue
                col_name = header[i].strip()
                val = cells[i].strip() if i < len(cells) else ""
                if not col_name or not val:
                    continue
                lines.append(f"- **{col_name}**: {val}")
            lines.append("")

        if len(lines) == 1:
            lines.append("*空数据*\n")

    return "\n".join(lines)


def process_excel(xlsx_path: str, output_dir: str, sub_dir: str = "") -> tuple[list[dict], list[str]]:
    """处理单个 Excel 文件，返回 (索引条目列表, 跳过的 sheet 名列表)。
    
    sub_dir: Excel 相对于输入根目录的子目录路径，用于在输出中保持层级。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    excel_base = Path(xlsx_path).stem  # 不含扩展名
    idx = []
    skipped = []

    # 在输出目录中创建对应的子目录
    target_dir = os.path.join(output_dir, sub_dir) if sub_dir else output_dir
    os.makedirs(target_dir, exist_ok=True)

    for ws in wb.worksheets:
        sheet_name = ws.title
        if _is_header_sheet(sheet_name):
            skipped.append(sheet_name)
            continue

        md_name = f"{safe_path(excel_base)}__{safe_path(sheet_name)}.md"
        md_path = os.path.join(target_dir, md_name)

        md_content = table_to_md(ws)
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(md_content)

        idx.append({
            "path": os.path.relpath(md_path, output_dir),
            "excel": excel_base,
            "sheet": sheet_name,
        })

    wb.close()
    return idx, skipped


def write_index(entries: list[dict], version: str, output_dir: str):
    """写入 INDEX.md。"""
    lines = [
        "# Excel → Markdown 索引",
        "",
        f"**版本**: {version}",
        "",
        "| 路径 | Excel 名称 | Sheet 名称 |",
        "|------|-----------|-----------|",
    ]
    for e in entries:
        lines.append(f"| `{e['path']}` | {e['excel']} | {e['sheet']} |")
    lines.append("")

    with open(os.path.join(output_dir, "INDEX.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def find_excel_files(root: str) -> tuple[list[str], list[str]]:
    """递归查找所有 Excel 文件。
    返回 (xlsx_list, xls_list)，.xls 因 openpyxl 不支持需单独处理。
    """
    xlsx_files = []
    xls_files = []
    for dirpath, _, filenames in os.walk(root):
        for fn in sorted(filenames):
            ext = Path(fn).suffix.lower()
            if ext == ".xlsx":
                xlsx_files.append(os.path.join(dirpath, fn))
            elif ext == ".xls":
                xls_files.append(os.path.join(dirpath, fn))
    return sorted(xlsx_files), sorted(xls_files)


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input_dir", help="包含 Excel 文件的目录")
    parser.add_argument("--output", "-o", help="输出目录（默认: <input_dir>/excel_md_output）")
    args = parser.parse_args()

    root = os.path.abspath(args.input_dir)
    if not os.path.isdir(root):
        sys.exit(f"错误: 目录不存在: {root}")

    out = args.output or os.path.join(root, "excel_md_output")
    os.makedirs(out, exist_ok=True)

    version = extract_version(root)
    xlsx_files, xls_files = find_excel_files(root)
    if not xlsx_files and not xls_files:
        print(f"未在 {root} 下找到任何 Excel 文件")
        return

    if xls_files:
        print(f"⚠ 跳过 {len(xls_files)} 个 .xls 文件 (openpyxl 不支持旧格式):")
        for f in xls_files:
            print(f"    {os.path.relpath(f, root)}")
        print()

    print(f"版本: {version}")
    print(f"找到 {len(xlsx_files)} 个 .xlsx 文件")
    print(f"输出目录: {out}\n")

    all_entries = []
    for fp in xlsx_files:
        rel = os.path.relpath(fp, root)
        sub_dir = safe_path(str(Path(rel).parent))  # Excel 所在子目录，空格→_
        print(f"  处理: {rel}")
        entries, skipped = process_excel(fp, out, sub_dir)
        if skipped:
            print(f"    跳过页头: {', '.join(skipped)}")
        all_entries.extend(entries)

    write_index(all_entries, version, out)
    print(f"\n完成! 共生成 {len(all_entries)} 个 md 文件")
    print(f"索引文件: {os.path.join(out, 'INDEX.md')}")


if __name__ == "__main__":
    main()
